"""Сборка Excel-отчёта ProcessFix AI через openpyxl.

Книга содержит 3 листа:
  1. «Пульс»        — сводные KPI за сутки
  2. «Потери ФОТ»   — детальная таблица по операциям
  3. «AI Анализ»    — гипотезы «5 Почему» от LLM
"""

from __future__ import annotations

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.analytics import DailyReport

logger = logging.getLogger(__name__)

# ── Стили ───────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=11, color="595959")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=28, color="C00000")
KPI_LABEL_FONT = Font(name="Calibri", bold=False, size=12, color="404040")

MONEY_FORMAT = '#,##0.00 "₽"'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _build_pulse_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Лист «Пульс» — сводные метрики."""
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "ProcessFix AI — Пульс дня"
    title_cell.font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Отчёт за: {report.report_date:%d.%m.%Y %H:%M} UTC"
    ws["A2"].font = SUBTITLE_FONT

    ws["A4"].value = "Суммарные потери ФОТ"
    ws["A4"].font = KPI_LABEL_FONT
    ws["A5"].value = report.total_loss_rub
    ws["A5"].font = KPI_VALUE_FONT
    ws["A5"].number_format = MONEY_FORMAT

    ws["C4"].value = "Операций обработано"
    ws["C4"].font = KPI_LABEL_FONT
    ws["C5"].value = report.total_events
    ws["C5"].font = Font(name="Calibri", bold=True, size=28, color="2F5496")

    anomaly_count = len(report.top_anomalies)
    ws["A7"].value = "Аномалий обнаружено"
    ws["A7"].font = KPI_LABEL_FONT
    ws["A8"].value = anomaly_count
    ws["A8"].font = Font(name="Calibri", bold=True, size=28, color="ED7D31")

    if report.top_anomalies:
        ws["A10"].value = "Топ-аномалия"
        ws["A10"].font = KPI_LABEL_FONT
        ws["A11"].value = report.top_anomalies[0].operation_name
        ws["A11"].font = Font(name="Calibri", bold=True, size=14, color="C00000")
        ws["A12"].value = f"Потери: {report.top_anomalies[0].total_loss_rub:,.2f} ₽"
        ws["A12"].font = SUBTITLE_FONT

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25


def _build_losses_sheet(ws: Worksheet, report: DailyReport) -> None:
    """Лист «Потери ФОТ» — детальная таблица."""
    ws.sheet_properties.tabColor = "C00000"

    headers = ["Операция", "Факт (ср. мин)", "Норма (мин)", "Δ (сек)", "Ставка (₽/ч)", "Событий", "Сумма потерь (₽)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    df = report.losses_df
    if df.empty:
        ws.cell(row=2, column=1, value="Нет данных")
        return

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=1, value=row["operation_name"])
        ws.cell(row=i, column=2, value=row["avg_duration_min"])
        ws.cell(row=i, column=3, value=row["norm_min"])
        ws.cell(row=i, column=4, value=round(row["delta_sec"], 1))
        ws.cell(row=i, column=5, value=row["hourly_rate_rub"])
        ws.cell(row=i, column=6, value=int(row["event_count"]))

        loss_cell = ws.cell(row=i, column=7, value=row["total_loss_rub"])
        loss_cell.number_format = MONEY_FORMAT

        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _build_ai_sheet(ws: Worksheet, report: DailyReport, ai_text: str) -> None:
    """Лист «AI Анализ» — гипотезы «5 Почему»."""
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:E1")
    ws["A1"].value = "AI-анализ корневых причин (метод «5 Почему»)"
    ws["A1"].font = TITLE_FONT

    if not report.top_anomalies:
        ws["A3"].value = "Аномалий не обнаружено — анализ не требуется."
        ws["A3"].font = SUBTITLE_FONT
        return

    top = report.top_anomalies[0]
    ws["A3"].value = f"Операция: {top.operation_name}"
    ws["A3"].font = Font(name="Calibri", bold=True, size=12)
    ws["A4"].value = (
        f"Факт: {top.avg_duration_sec:.0f} сек | "
        f"Норма: {top.norm_seconds} сек | "
        f"Потери: {top.total_loss_rub:,.2f} ₽"
    )
    ws["A4"].font = SUBTITLE_FONT

    ws["A6"].value = "Гипотезы LLM:"
    ws["A6"].font = Font(name="Calibri", bold=True, size=11)

    for row_idx, line in enumerate(ai_text.split("\n"), start=7):
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 90


def build_excel_report(report: DailyReport, ai_text: str) -> bytes:
    """Собрать Excel-файл и вернуть его содержимое как bytes."""
    wb = Workbook()

    ws_pulse = wb.active
    ws_pulse.title = "Пульс"
    _build_pulse_sheet(ws_pulse, report)

    ws_losses = wb.create_sheet("Потери ФОТ")
    _build_losses_sheet(ws_losses, report)

    ws_ai = wb.create_sheet("AI Анализ")
    _build_ai_sheet(ws_ai, report, ai_text)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    logger.info("Excel-отчёт сгенерирован: %.1f KB", len(file_bytes) / 1024)
    return file_bytes
