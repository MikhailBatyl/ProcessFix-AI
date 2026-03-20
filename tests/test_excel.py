"""Unit-тесты для app.services.excel — генерация xlsx-отчёта."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from app.services.analytics import DailyReport
from app.services.excel import build_excel_report


class TestBuildExcelReport:

    def test_returns_bytes(self, sample_report: DailyReport):
        result = build_excel_report(sample_report, "Test AI text")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_structure(self, sample_report: DailyReport):
        file_bytes = build_excel_report(sample_report, "1. Гипотеза A\n2. Гипотеза B")
        wb = load_workbook(io.BytesIO(file_bytes))

        assert "Пульс" in wb.sheetnames
        assert "Потери ФОТ" in wb.sheetnames
        assert "AI Анализ" in wb.sheetnames
        assert len(wb.sheetnames) == 3

    def test_pulse_sheet_has_kpi_values(self, sample_report: DailyReport):
        file_bytes = build_excel_report(sample_report, "AI text")
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb["Пульс"]

        assert ws["A1"].value == "ProcessFix AI — Пульс дня"
        assert ws["A5"].value == sample_report.total_loss_rub
        assert ws["C5"].value == sample_report.total_events

    def test_losses_sheet_has_header_and_data(self, sample_report: DailyReport):
        file_bytes = build_excel_report(sample_report, "")
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb["Потери ФОТ"]

        assert ws.cell(row=1, column=1).value == "Операция"
        assert ws.cell(row=1, column=7).value == "Сумма потерь (₽)"
        assert ws.cell(row=2, column=1).value is not None

    def test_ai_sheet_contains_text(self, sample_report: DailyReport):
        ai_text = "1. Первая гипотеза\n2. Вторая гипотеза"
        file_bytes = build_excel_report(sample_report, ai_text)
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb["AI Анализ"]

        assert ws["A1"].value == "AI-анализ корневых причин (метод «5 Почему»)"
        values = [ws.cell(row=r, column=1).value for r in range(7, 10)]
        found = any("Первая гипотеза" in str(v) for v in values if v)
        assert found

    def test_empty_report_still_generates(self):
        import pandas as pd
        empty_report = DailyReport(
            report_date=__import__("datetime").datetime(2026, 1, 1),
            total_loss_rub=0.0,
            total_events=0,
            losses_df=pd.DataFrame(),
            top_anomalies=[],
        )
        file_bytes = build_excel_report(empty_report, "")
        assert isinstance(file_bytes, bytes)

        wb = load_workbook(io.BytesIO(file_bytes))
        assert "Пульс" in wb.sheetnames
